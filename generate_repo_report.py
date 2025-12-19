import argparse
import datetime as dt
import os
import subprocess
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


EXCLUDE_DIRS = {
    ".git", ".vs", "bin", "obj", ".idea", ".vscode", "node_modules",
    "__pycache__", ".pytest_cache", "dist", "build", "out"
}
BINARY_EXTS = {
    ".png",".jpg",".jpeg",".gif",".webp",".ico",".pdf",".dll",".exe",".pdb",".db",
    ".bin",".zip",".7z",".rar",".mp4",".mp3",".wav",".woff",".woff2",".ttf",".otf"
}

LANG_BY_EXT = {
    ".cs": "C#",
    ".cshtml": "Razor (CSHTML)",
    ".razor": "Razor",
    ".html": "HTML",
    ".css": "CSS",
    ".js": "JavaScript",
    ".ts": "TypeScript",
    ".json": "JSON",
    ".xml": "XML",
    ".yml": "YAML",
    ".yaml": "YAML",
    ".md": "Markdown",
    ".sln": "Solution",
    ".csproj": "C# Project",
    ".props": "MSBuild Props",
    ".targets": "MSBuild Targets",
    ".config": "Config",
}

def run(cmd, cwd=None):
    p = subprocess.run(cmd, cwd=cwd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
    if p.returncode != 0:
        raise RuntimeError(f"Command failed: {' '.join(cmd)}\n{p.stderr.strip()}")
    return p.stdout

def week_start_monday(d: dt.date) -> dt.date:
    return d - dt.timedelta(days=d.weekday())

def parse_date(s: str) -> dt.date:
    return dt.date.fromisoformat(s)

def ensure_repo(repo_url: str | None, repo_path: str | None) -> str:
    if repo_path:
        rp = os.path.abspath(repo_path)
        if not os.path.isdir(os.path.join(rp, ".git")):
            raise RuntimeError(f"repo-path не похож на git-репозиторий: {rp}")
        return rp

    # clone
    if not repo_url:
        raise RuntimeError("Нужно указать --repo-url или --repo-path")

    base = os.path.abspath("./_repo_tmp")
    os.makedirs(base, exist_ok=True)
    target = os.path.join(base, "repo")
    if os.path.isdir(os.path.join(target, ".git")):
        # update
        run(["git", "fetch", "--all", "--prune"], cwd=target)
    else:
        if os.path.exists(target):
            # remove old junk
            for root, dirs, files in os.walk(target, topdown=False):
                for f in files:
                    try: os.remove(os.path.join(root, f))
                    except: pass
                for d in dirs:
                    try: os.rmdir(os.path.join(root, d))
                    except: pass
            try: os.rmdir(target)
            except: pass
        run(["git", "clone", "--depth", "999999", repo_url, target], cwd=base)
    return target

def commits_per_week(repo_dir: str, start: dt.date, end: dt.date) -> Counter:
    # ISO date like "2025-12-19 12:34:56 +0300"
    out = run(["git", "log", "--date=iso-strict", "--pretty=%ad"], cwd=repo_dir)
    c = Counter()
    for line in out.splitlines():
        line = line.strip()
        if not line:
            continue
        # iso-strict example: 2025-12-19T10:11:12+03:00
        d = dt.datetime.fromisoformat(line).date()
        wk = week_start_monday(d)
        if start <= wk <= end:
            c[wk] += 1
    return c

def walk_project(repo_dir: str):
    total_files = 0
    total_bytes = 0
    src_files = 0
    loc_total = 0
    loc_by_lang = Counter()
    files_by_lang = Counter()
    bytes_by_lang = Counter()

    for root, dirs, files in os.walk(repo_dir):
        # prune excluded dirs
        dirs[:] = [d for d in dirs if d.lower() not in EXCLUDE_DIRS]
        for f in files:
            total_files += 1
            p = os.path.join(root, f)
            try:
                st = os.stat(p)
            except:
                continue
            total_bytes += st.st_size

            ext = Path(f).suffix.lower()
            if ext in BINARY_EXTS:
                continue

            # treat as source/text
            src_files += 1
            lang = LANG_BY_EXT.get(ext, ext.upper() if ext else "Other")
            bytes_by_lang[lang] += st.st_size
            files_by_lang[lang] += 1

            # LOC
            try:
                with open(p, "rb") as fh:
                    b = fh.read()
                try:
                    text = b.decode("utf-8")
                except UnicodeDecodeError:
                    text = b.decode("latin-1", errors="replace")
                loc = text.count("\n") + (1 if text else 0)
            except:
                loc = 0
            loc_total += loc
            loc_by_lang[lang] += loc

    return {
        "total_files": total_files,
        "total_bytes": total_bytes,
        "src_files": src_files,
        "loc_total": loc_total,
        "loc_by_lang": loc_by_lang,
        "files_by_lang": files_by_lang,
        "bytes_by_lang": bytes_by_lang,
    }

def build_weeks(start: dt.date, end: dt.date):
    w = week_start_monday(start)
    endw = week_start_monday(end)
    weeks = []
    while w <= endw:
        weeks.append(w)
        w += dt.timedelta(days=7)
    return weeks

def style_header(row):
    for cell in row:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E6EEF8")
        cell.alignment = Alignment(horizontal="center")

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--repo-url", default=None)
    ap.add_argument("--repo-path", default=None)
    ap.add_argument("--out", required=True)
    ap.add_argument("--start", required=True, help="YYYY-MM-DD (например 2025-09-01)")
    ap.add_argument("--end", required=True, help="YYYY-MM-DD (например 2025-12-31)")
    args = ap.parse_args()

    start = parse_date(args.start)
    end = parse_date(args.end)
    weeks = build_weeks(start, end)

    repo_dir = ensure_repo(args.repo_url, args.repo_path)

    weekly_commits = commits_per_week(repo_dir, week_start_monday(start), week_start_monday(end))
    stats = walk_project(repo_dir)

    wb = Workbook()

    # Sheet 1: weekly activity
    ws = wb.active
    ws.title = "Активность (недели)"
    ws["A1"] = "Активность по написанию кода (коммиты в git по неделям)"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:C1")

    ws.append(["Неделя (пн)", "Коммитов (шт)", "Комментарий"])
    style_header(ws[2])

    for wk in weeks:
        ws.append([wk.isoformat(), int(weekly_commits[wk]), ""])

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 40
    ws.freeze_panes = "A3"

    chart = BarChart()
    chart.title = "Коммиты по неделям"
    chart.y_axis.title = "Коммиты"
    chart.x_axis.title = "Неделя"
    data = Reference(ws, min_col=2, min_row=2, max_row=2+len(weeks))
    cats = Reference(ws, min_col=1, min_row=3, max_row=2+len(weeks))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 24
    ws.add_chart(chart, "E2")

    # Sheet 2: project size
    ws2 = wb.create_sheet("Объем проекта")
    ws2["A1"] = "Объем проекта"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2.merge_cells("A1:D1")

    ws2.append(["Метрика", "Значение"])
    style_header(ws2[2])

    ws2.append(["Всего файлов (включая всё)", stats["total_files"]])
    ws2.append(["Размер репозитория (MB)", round(stats["total_bytes"]/1024/1024, 2)])
    ws2.append(["Файлов текста/кода (без bin/obj/.vs/...)", stats["src_files"]])
    ws2.append(["LOC (строк)", stats["loc_total"]])

    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 18

    ws2.append([])
    ws2.append(["Разбивка по типам", "", "", ""])
    ws2["A8"].font = Font(bold=True)

    ws2.append(["Язык/тип", "Файлов", "LOC", "Размер (KB)"])
    style_header(ws2[9])

    # sort by LOC desc
    langs = sorted(stats["loc_by_lang"].items(), key=lambda x: x[1], reverse=True)
    for lang, loc in langs:
        ws2.append([
            lang,
            int(stats["files_by_lang"][lang]),
            int(loc),
            round(stats["bytes_by_lang"][lang]/1024, 1)
        ])

    # Pie chart for LOC (top 8 + others)
    loc_items = [(k,v) for k,v in stats["loc_by_lang"].items() if v > 0]
    loc_items.sort(key=lambda x: x[1], reverse=True)
    top = loc_items[:8]
    others = sum(v for _,v in loc_items[8:])
    pie_rows = top + ([("Others", others)] if others > 0 else [])

    base_row = 9
    base_col = 6
    ws2.cell(row=base_row, column=base_col, value="Категория").font = Font(bold=True)
    ws2.cell(row=base_row, column=base_col+1, value="LOC").font = Font(bold=True)
    r = base_row + 1
    for k,v in pie_rows:
        ws2.cell(row=r, column=base_col, value=k)
        ws2.cell(row=r, column=base_col+1, value=int(v))
        r += 1

    pie = PieChart()
    pie.title = "Доля LOC по типам"
    labels = Reference(ws2, min_col=base_col, min_row=base_row+1, max_row=r-1)
    data = Reference(ws2, min_col=base_col+1, min_row=base_row, max_row=r-1)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.height = 12
    pie.width = 18
    ws2.add_chart(pie, "F2")

    wb.save(args.out)
    print(f"OK: сохранено в {args.out}")

if __name__ == "__main__":
    main()
